import os
import re
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

# === CONFIGURATION ===

excel_folder = 'Rack/Rack Vendors'             # Folder containing Excel files
output_folder = 'Rack/Rack_images'             # Folder to save all extracted images
name_column_index = 1                              # 1 = Column A, where "Name" is

# === PREPARE OUTPUT DIRECTORY ===
os.makedirs(output_folder, exist_ok=True)

# === PROCESS EACH EXCEL FILE IN FOLDER ===
for filename in os.listdir(excel_folder):
    if filename.endswith('.xlsx') and not filename.startswith('~$'):  # Skip temp Excel files
        excel_path = os.path.join(excel_folder, filename)
        print(f"🔍 Processing: {excel_path}")

        try:
            wb = load_workbook(excel_path)
            ws = wb.active

            # Build mapping: Row -> Name
            row_to_name = {}
            for row in ws.iter_rows(min_row=2):
                name_cell = row[name_column_index - 1]
                if name_cell.value:
                    row_to_name[name_cell.row] = str(name_cell.value)

            # Extract images
            for image in ws._images:
                if hasattr(image, 'anchor'):
                    row = image.anchor._from.row + 1
                    name = row_to_name.get(row)

                    if name:
                        try:
                            # Clean product name
                            safe_name = re.sub(r'[\\/*?:"<>|\n\r]+', "_", name).strip()
                            output_filename = f"{safe_name}.png"
                            output_path = os.path.join(output_folder, output_filename)

                            # Save image
                            img_data = image._data()
                            img = Image.open(BytesIO(img_data))
                            img.save(output_path)

                            print(f"✅ Saved: {output_path}")

                        except Exception as e:
                            print(f"❌ Error saving image for row {row} ({name}): {e}")
                    else:
                        print(f"⚠️ No name found for image at row {row}")

        except Exception as e:
            print(f"❌ Failed to process {filename}: {e}")

print("✅ All files processed.")
