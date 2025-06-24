

import os
import re
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

# === CONFIGURATION ===
excel_file = 'Ports/original_Ports.xlsx'      # Path to your Excel file
output_folder = 'Ports/Ports_images'            # Folder to save extracted images
name_column_index = 1                       # 1 = Column A, where "Name" is

# === PREPARE OUTPUT DIRECTORY ===
os.makedirs(output_folder, exist_ok=True)

# === LOAD EXCEL WORKBOOK ===
wb = load_workbook(excel_file)
ws = wb.active

# === BUILD ROW -> NAME MAPPING ===
row_to_name = {}
for row in ws.iter_rows(min_row=2):  # Skip header row
    name_cell = row[name_column_index - 1]
    if name_cell.value:
        row_to_name[name_cell.row] = str(name_cell.value)

# === EXTRACT IMAGES ===
for image in ws._images:
    if hasattr(image, 'anchor'):
        row = image.anchor._from.row + 1  # openpyxl uses 0-based index
        name = row_to_name.get(row)

        if name:
            try:
                # Clean invalid characters from filename
                safe_name = re.sub(r'[\\/*?:"<>|\n\r]+', "_", name).strip()
                output_path = os.path.join(output_folder, f"{safe_name}.png")

                # Convert image bytes and save
                img_data = image._data()
                img = Image.open(BytesIO(img_data))
                img.save(output_path)

                print(f"✅ Saved: {output_path}")
            except Exception as e:
                print(f"❌ Error saving image for row {row} ({name}): {e}")
        else:
            print(f"⚠️ No name found for image at row {row}")

print("✅ All done.")
